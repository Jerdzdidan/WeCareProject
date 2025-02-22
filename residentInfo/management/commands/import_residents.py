import openpyxl
from django.core.management.base import BaseCommand
from datetime import datetime, date
from residentInfo.models import Family, Resident

# Map for category conversions from your "HOUSEHOLD INFORMATION (PRIORITY SECTORS)" column
CATEGORY_MAP = {
    "NOT APPLICABLE": "N/A",
    "CHILDREN": "Children",
    "SENIOR CITIZEN": "Senior",
    "SOLO PARENT": "Solo Parent",
    "PREGNANT": "Pregnant",
    "PWD": "PWD"
}

# Map for short gender codes to full model choices
GENDER_MAP = {
    "M": "Male",
    "F": "Female"
}

class Command(BaseCommand):
    help = "Import residents from an Excel file matching the provided columns."

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help="Path to the Excel file containing resident information."
        )

    def split_name(self, full_name):
        """
        Splits a string like 'GARCIA, JUAN DELA CRUZ' into:
          last_name='GARCIA', first_name='JUAN', middle_name='DELA CRUZ'
        If parsing fails, defaults to last_name=full_name, first_name='--', middle_name='--'
        """
        full_name = (full_name or "").strip()
        if not full_name:
            return "--", "--", "--"

        # Expecting format: LASTNAME, FIRSTNAME [MIDDLE...]
        parts = full_name.split(',')
        if len(parts) == 2:
            last_name = parts[0].strip()
            remainder = parts[1].strip()
            subparts = remainder.split()
            if len(subparts) == 1:
                # e.g. "GARCIA, JUAN"
                return last_name, subparts[0], "--"
            else:
                # e.g. "GARCIA, JUAN DELA CRUZ"
                return last_name, subparts[0], " ".join(subparts[1:])
        else:
            # Fallback if no comma or unexpected format
            return full_name, "--", "--"

    def parse_birthdate(self, birthdate_str):
        """
        Parse a birthdate in 'MM/DD/YYYY' format. If parsing fails, returns 2000-01-01.
        """
        if not birthdate_str:
            return date(2000, 1, 1)
        birthdate_str = str(birthdate_str).strip()
        try:
            # Example format: 3/15/1980
            return datetime.strptime(birthdate_str, "%m/%d/%Y").date()
        except ValueError:
            return date(2000, 1, 1)

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        try:
            workbook = openpyxl.load_workbook(excel_file)
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Error opening file: {e}"))
            return

        sheet = workbook.active

        # Expecting these 12 columns in the first row:
        expected_header = [
            "FAMILY NO.", "RESIDENT ID", "NAME", "RELATIONSHIP TO HEAD OF THE FAMILY",
            "BIRTHDATE", "AGE", "CIVIL STATUS", "GENDER",
            "HOUSEHOLD INFORMATION (PRIORITY SECTORS)", "PRESENT ADDRESS",
            "CONTACT NUMBER", "DATE UPDATED"
        ]
        # Read the header row
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_row = [str(cell).strip() if cell else "" for cell in header_row]

        if header_row != expected_header:
            self.stderr.write("Header does not match expected format.")
            self.stderr.write(f"Found header: {header_row}")
            self.stderr.write(f"Expected header: {expected_header}")
            return

        row_count = 0
        # Iterate over data rows, starting from row 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                family_no, resident_id, full_name, relationship,
                birthdate_str, age_val, civil_status,
                gender_val, household_info, present_address,
                contact_number, date_updated
            ) = row

            # 1. Construct Family primary key => "F" + family_no
            if family_no:
                family_no_str = str(family_no).strip()
                if not family_no_str.startswith("F"):
                    family_no_str = "F" + family_no_str
            else:
                family_no_str = None

            # 2. Get or create Family
            if family_no_str:
                family_obj, _ = Family.objects.get_or_create(family_no=family_no_str)
            else:
                family_obj = Family.objects.create()

            # 3. Construct Resident primary key => "R" + resident_id
            #    If resident_id is missing or invalid, we'll let the model auto-generate.
            resident_pk = None
            if resident_id:
                resident_pk = "R" + str(resident_id).strip()

            # 4. Parse NAME into last_name, first_name, middle_name
            last_name, first_name, middle_name = self.split_name(full_name)

            # 5. Relationship
            relationship = (relationship or "").strip() or "--"

            # 6. Birthdate
            birthdate = self.parse_birthdate(birthdate_str)

            # 7. Age
            try:
                age = int(age_val)
            except (ValueError, TypeError):
                age = 0

            # 8. Civil Status
            civil_status = (civil_status or "").strip() or "--"

            # 9. Gender => map "M" -> "Male", "F" -> "Female", else "Other"
            gender_val = (gender_val or "").strip().upper()
            gender = GENDER_MAP.get(gender_val, "Other")

            # 10. Category => from "household_info"
            household_str = (household_info or "").strip().upper()
            # Make uppercase for easy matching: "NOT APPLICABLE" -> "N/A", etc.
            category = CATEGORY_MAP.get(household_str, "N/A")

            # 11. Present Address
            present_address = (present_address or "").strip() or "--"

            # 12. Contact Number
            contact_number = str(contact_number or "").strip() or "--"


            # Create or update the Resident record
            # If we pass 'id=res_pk' and it doesn't exist, it will create. If it does exist, it might error or you can do update_or_create.
            if resident_pk:
                # We'll do an update_or_create so that if it already exists, we update it.
                resident, _ = Resident.objects.update_or_create(
                    id=resident_pk,
                    defaults={
                        'family': family_obj,
                        'last_name': last_name,
                        'first_name': first_name,
                        'middle_name': middle_name,
                        'relationship_to_head': relationship,
                        'birthdate': birthdate,
                        'age': age,
                        'civil_status': civil_status,
                        'gender': gender,
                        'category': category,
                        'present_address': present_address,
                        'contact_number': contact_number,
                    }
                )
            else:
                # Let the model auto-generate the resident ID
                resident = Resident.objects.create(
                    family=family_obj,
                    last_name=last_name,
                    first_name=first_name,
                    middle_name=middle_name,
                    relationship_to_head=relationship,
                    birthdate=birthdate,
                    age=age,
                    civil_status=civil_status,
                    gender=gender,
                    category=category,
                    present_address=present_address,
                    contact_number=contact_number,
                )

            row_count += 1
            self.stdout.write(self.style.SUCCESS(
                f"Imported Resident {resident.id}: {resident.first_name} {resident.last_name} (Family: {family_obj.family_no})"
            ))

        self.stdout.write(self.style.SUCCESS(f"Import complete. Total rows processed: {row_count}"))
