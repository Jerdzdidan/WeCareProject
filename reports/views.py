from django.shortcuts import render
from residentInfo.models import Resident
from patientInfo.models import Patient
from django.http import HttpResponse
from django.template.loader import render_to_string
import io
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill
from django.utils import timezone
from xhtml2pdf import pisa
from users.decorators import role_required
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from datetime import date
from medicineMonitoring.models import Medicine


# Create your views here.
def residentInfoReport(request):
    category = request.GET.get('category', '')
    gender = request.GET.get('gender', '')
    age_group = request.GET.get('age', '')

    residents = Resident.objects.select_related('family').all()

    if category:
        residents = residents.filter(category=category)
    if gender:
        residents = residents.filter(gender=gender)
    
    if age_group:
        try:
            age_group = int(age_group)
        except ValueError:
            age_group = None
        
        if age_group is not None:
            if age_group == 6:
                residents = residents.filter(age__lte=6)
            elif age_group == 18:
                residents = residents.filter(age__gte=7, age__lte=18)
            elif age_group == 59:
                residents = residents.filter(age__gte=19, age__lte=59)
            elif age_group == 150:
                residents = residents.filter(age__gte=60)

    residents = residents.order_by('family__family_no', 'id')

    return render(request, 'reports/residentInfoReport.html', {'residents': residents})

# Printing and exporting of residents
def get_filtered_residents(request):
    residents = Resident.objects.all()

    category = request.GET.get('category')
    if category:
        residents = residents.filter(category=category)
    gender = request.GET.get('gender')
    if gender:
        residents = residents.filter(gender=gender)
    age_filter = request.GET.get('age')
    if age_filter:
        try:
            age_value = int(age_filter)
            if age_value == 6:
                residents = residents.filter(age__lte=6)
            elif age_value == 18:
                residents = residents.filter(age__gte=7, age__lte=18)
            elif age_value == 59:
                residents = residents.filter(age__gte=19, age__lte=59)
            elif age_value == 150:
                residents = residents.filter(age__gte=60)
        except ValueError:
            pass
    return residents


def resident_export_xlsx(request):
    residents = get_filtered_residents(request)
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Residents Report"

    # Title styling (row 1)
    title_font = Font(bold=True, size=14)
    ws.merge_cells('A1:N1')
    ws['A1'] = "WeCare - A Resident and Patient Management System"
    ws['A1'].font = title_font

    ws.merge_cells('A2:N2')
    ws['A2'] = f"Category: {request.GET.get('category', 'All')}"
    
    ws.merge_cells('A3:N3')
    ws['A3'] = f"Gender: {request.GET.get('gender', 'All')}"
    
    ws.merge_cells('A4:N4')
    ws['A4'] = f"Age-group: {request.GET.get('age', 'All')}"
    
    headers = [
        "Family No.", "Resident ID", "Last Name", "First Name", "Middle Name",
        "Relationship to Head", "Birthdate", "Age", "Civil Status", "Gender",
        "Category", "Address", "Contact Number", "Date Updated"
    ]
    header_font = Font(bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    current_row = 7
    for resident in residents:
        ws.cell(row=current_row, column=1, value=resident.family.family_no if resident.family else "")
        ws.cell(row=current_row, column=2, value=resident.id)
        ws.cell(row=current_row, column=3, value=resident.last_name)
        ws.cell(row=current_row, column=4, value=resident.first_name)
        ws.cell(row=current_row, column=5, value=resident.middle_name)
        ws.cell(row=current_row, column=6, value=resident.relationship_to_head)
        ws.cell(row=current_row, column=7, value=resident.birthdate.strftime("%b. %d, %Y") if resident.birthdate else "")
        ws.cell(row=current_row, column=8, value=resident.age)
        ws.cell(row=current_row, column=9, value=resident.civil_status)
        ws.cell(row=current_row, column=10, value=resident.gender)
        ws.cell(row=current_row, column=11, value=resident.category)
        ws.cell(row=current_row, column=12, value=resident.present_address)
        ws.cell(row=current_row, column=13, value=resident.contact_number)
        ws.cell(row=current_row, column=14, value=resident.date_updated.strftime("%b. %d, %Y") if resident.date_updated else "")
        current_row += 1

    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output, 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="WeCare Residents Report.xlsx"'
    return response


def resident_export_pdf(request):
    residents = get_filtered_residents(request)
    context = {
        'residents': residents,
        'filter_category': request.GET.get('category', 'All'),
        'filter_gender': request.GET.get('gender', 'All'),
        'filter_age': request.GET.get('age', 'All'),
        'printed_by': request.user.get_full_name() if request.user.is_authenticated else 'Anonymous',
        'printed_on': timezone.now(),
    }
    html = render_to_string('reports/resident_pdf_template.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="WeCare Residents Report.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF <pre>' + html + '</pre>')
    
    return response


# Patient Report
@login_required
def patientInfoReport(request):
    category = request.GET.get('category', '')
    gender = request.GET.get('gender', '')
    street = request.GET.get('street', '')
    age_group = request.GET.get('age', '')  

    patients = Patient.objects.select_related('resident').all()

    if category:
        patients = patients.filter(resident__category=category)
    if gender:
        patients = patients.filter(resident__gender=gender)
    if street:
        patients = patients.filter(resident__present_address__icontains=street)
    
    if age_group:
        try:
            age_group = int(age_group)
        except ValueError:
            age_group = None
        
        if age_group is not None:
            if age_group == 6:
                patients = patients.filter(resident__age__lte=6)
            elif age_group == 18:
                patients = patients.filter(resident__age__gte=7, resident__age__lte=18)
            elif age_group == 59:
                patients = patients.filter(resident__age__gte=19, resident__age__lte=59)
            elif age_group == 150:
                patients = patients.filter(resident__age__gte=60)
    
    patients = patients.order_by('patientID')

    return render(request, 'reports/patientInfoReport.html', {'patients': patients})

# Add these patient export functions to your views.py

def get_filtered_patients(request):
    patients = Patient.objects.select_related('resident').all()

    category = request.GET.get('category')
    if category:
        patients = patients.filter(resident__category=category)
    
    gender = request.GET.get('gender')
    if gender:
        patients = patients.filter(resident__gender=gender)
    
    street = request.GET.get('street')
    if street:
        patients = patients.filter(resident__present_address__icontains=street)
    
    age_filter = request.GET.get('age')
    if age_filter:
        try:
            age_value = int(age_filter)
            if age_value == 6:
                patients = patients.filter(resident__age__lte=6)
            elif age_value == 18:
                patients = patients.filter(resident__age__gte=7, resident__age__lte=18)
            elif age_value == 59:
                patients = patients.filter(resident__age__gte=19, resident__age__lte=59)
            elif age_value == 150:
                patients = patients.filter(resident__age__gte=60)
        except ValueError:
            pass
    
    return patients.order_by('patientID')

@login_required
def patient_export_xlsx(request):
    patients = get_filtered_patients(request)
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients Report"

    title_font = Font(bold=True, size=14)
    ws.merge_cells('A1:L1')
    ws['A1'] = "WeCare - Patient Medical Report"
    ws['A1'].font = title_font

    ws.merge_cells('A2:L2')
    ws['A2'] = f"Category: {request.GET.get('category', 'All')}"
    ws.merge_cells('A3:L3')
    ws['A3'] = f"Gender: {request.GET.get('gender', 'All')}"
    ws.merge_cells('A4:L4')
    ws['A4'] = f"Age-group: {request.GET.get('age', 'All')}"

    headers = [
        "Family No.",
        "Patient ID",
        "Last Name",
        "First Name",
        "Middle Name",
        "Birthdate",
        "Age",
        "Civil Status",
        "Gender",
        "Category",
        "Contact Number",
        "Last Visit",
    ]
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    header_row = 6
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    current_row = header_row + 1
    for patient in patients:
        resident = patient.resident
        family_no = resident.family.family_no if hasattr(resident, 'family') and resident.family else ""
        ws.cell(row=current_row, column=1, value=family_no)
        ws.cell(row=current_row, column=2, value=patient.patientID)
        ws.cell(row=current_row, column=3, value=resident.last_name)
        ws.cell(row=current_row, column=4, value=resident.first_name)
        ws.cell(row=current_row, column=5, value=resident.middle_name)
        ws.cell(row=current_row, column=6, 
                value=resident.birthdate.strftime("%b %d, %Y") if resident.birthdate else "")
        ws.cell(row=current_row, column=7, value=resident.age)
        ws.cell(row=current_row, column=8, value=resident.civil_status)
        ws.cell(row=current_row, column=9, value=resident.gender)
        ws.cell(row=current_row, column=10, value=resident.category)
        ws.cell(row=current_row, column=11, value=resident.contact_number)
        ws.cell(row=current_row, column=12, 
                value=patient.last_visit_date.strftime("%b %d, %Y") if patient.last_visit_date else "--")
        current_row += 1

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output, 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="WeCare Patients Report.xlsx"'
    return response

@login_required
def patient_export_pdf(request):
    patients = get_filtered_patients(request)
    context = {
        'patients': patients,
        'filter_category': request.GET.get('category', 'All'),
        'filter_gender': request.GET.get('gender', 'All'),
        'filter_street': request.GET.get('street', 'All'),
        'filter_age': request.GET.get('age', 'All'),
        'printed_by': request.user.get_full_name() if request.user.is_authenticated else 'Anonymous',
        'printed_on': timezone.now(),
    }
    html = render_to_string('reports/patient_pdf_template.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="WeCare Patients Report.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF')
    
    return response



@login_required
@role_required(['ADMIN', 'BHW', 'DOCTOR'], 'Medicine Record')
def medicineReport(request):
    today = date.today()
    medicines = Medicine.objects.all().order_by("medicine_name").annotate(
        releasedQty=Sum('medicine_trackings__quantity_used'),
        expiredQty=Sum('stocks__quantity', filter=Q(stocks__expiration_date__lte=today))
    )
    
    context = {
        'medicines': medicines,
        'today': today,
    }
    return render(request, "reports/medicineReport.html", context)

@login_required
@role_required(['ADMIN', 'BHW', 'DOCTOR'], 'Medicine Record')
def medicine_export_xlsx(request):
    today = date.today()
    # Annotate each medicine with released and expired quantities.
    medicines = Medicine.objects.all().order_by("medicine_name").annotate(
        releasedQty=Sum('medicine_trackings__quantity_used'),
        expiredQty=Sum('stocks__quantity', filter=Q(stocks__expiration_date__lte=today))
    )
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medicines Report"

    # Title row
    title_font = Font(bold=True, size=14)
    ws.merge_cells('A1:L1')
    ws['A1'] = "WeCare - Medicine Report"
    ws['A1'].font = title_font

    # Printed date info (optional)
    ws.merge_cells('A2:L2')
    ws['A2'] = f"Printed on: {today.strftime('%b %d, %Y')}"

    # Header row
    headers = [
        "Medicine Name", "Dosage", "Generic Name", "Brand Name",
        "Unit Price", "Total Value", "Total Quantity",
        "Released Quantity", "Expired Quantity",
        "Supplier", "Date Last Stocked", "Notes"
    ]
    header_row = 4
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    current_row = header_row + 1
    for med in medicines:
        ws.cell(row=current_row, column=1, value=med.medicine_name)
        ws.cell(row=current_row, column=2, value=med.dosage)
        ws.cell(row=current_row, column=3, value=med.generic_name)
        ws.cell(row=current_row, column=4, value=med.brand_name)
        ws.cell(row=current_row, column=5, value=float(med.unit_price))
        ws.cell(row=current_row, column=6, value=float(med.total_value))
        ws.cell(row=current_row, column=7, value=med.total_quantity)
        ws.cell(row=current_row, column=8, value=med.releasedQty if med.releasedQty is not None else 0)
        ws.cell(row=current_row, column=9, value=med.expiredQty if med.expiredQty is not None else 0)
        ws.cell(row=current_row, column=10, value=med.supplier_name)
        ws.cell(row=current_row, column=11, value=med.date_last_stock.strftime("%b %d, %Y") if med.date_last_stock else "")
        ws.cell(row=current_row, column=12, value=med.notes)
        current_row += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="WeCare Medicines Report.xlsx"'
    return response


@login_required
@role_required(['ADMIN', 'BHW', 'DOCTOR'], 'Medicine Record')
def medicine_export_pdf(request):
    today = date.today()
    medicines = Medicine.objects.all().order_by("medicine_name").annotate(
        releasedQty=Sum('medicine_trackings__quantity_used'),
        expiredQty=Sum('stocks__quantity', filter=Q(stocks__expiration_date__lte=today))
    )
    context = {
        'medicines': medicines,
        'printed_on': timezone.now(),
        'today': today,
    }
    # Render the PDF template with the context.
    html = render_to_string('reports/medicine_pdf_template.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="WeCare Medicines Report.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Error generating PDF")
    return response
